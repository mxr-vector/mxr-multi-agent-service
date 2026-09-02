"""
文件摄取工具（解析 + 分块，纯预处理，不落库、不做向量化）。

职责边界（对应设计 D3）：
- 按文件类型（pdf/markdown/excel/docx/text/csv）把上传内容解析为完整 UTF-8 文本；
- 计算 content_hash = sha256(全文)，供增量同步判断源文件是否变更；
- 产出两级父子块树：level 1 父块 + level 0 叶块，叶块携带 parent 引用、
  每级顺序号 chunk_index、以及相对父块的 char_start/char_end 偏移；
- PDF 按页解析以便回填 page_start/page_end。

分块策略（用户可选，默认 char）：
- char：全文字符递归切分，chapter_title 恒为 None；
- structure：按章节/标题先切节，节内再字符切分，块携带 chapter_title；
- semantic：句子级 embedding + 相邻相似度阈值切语义父块，父块内再字符切叶块；
  embedding 服务异常时整体降级为 char 切分并记 warning。

该模块只依赖纯 Python 解析器与字符级切分器；语义分块经 EmbeddingFactory 走
HTTP embedding（符合"HTTP-only 模型推理"原则，不加载模型），失败自动降级。
业务层（service/rag/document.py）负责把产出持久化。
"""

import hashlib
import io
import re
from typing import Any

import numpy as np
from langchain_text_splitters import RecursiveCharacterTextSplitter

from exception.bad_except import BadException, bad_except
from model.embeddings.factory import EmbeddingFactory
from model.embeddings.similarity_utils import SimilarityUtils
from utils.logger import logger

# 两级切分窗口：level 1 父块给回写上下文，level 0 叶块作检索单元
PARENT_CHUNK_SIZE = 2000
CHILD_CHUNK_SIZE = 400
CHILD_CHUNK_OVERLAP = 80

# 用户可选切块策略：char 纯字符；structure 强制章节；semantic 语义相似度切分
CHUNK_STRATEGIES = ("char", "structure", "semantic")
DEFAULT_CHUNK_STRATEGY = "char"

# 语义切分参数：相邻句相似度低于阈值处断句；目标父块大小上限封口
SEMANTIC_SIMILARITY_THRESHOLD = 0.7
SEMANTIC_TARGET_SIZE = 2000

# 支持章节感知切分的 doc_type（pdf/text/csv 无可靠结构信号，本期不做）
_STRUCTURE_DOC_TYPES = {"markdown", "docx", "excel"}


def validate_chunk_strategy(strategy: str, doc_type: str) -> str:
    """
    校验策略取值与格式组合：非法取值、structure 搭配无结构格式均抛业务异常。
    semantic 对所有格式可用。返回校验通过的策略值。
    """
    if strategy not in CHUNK_STRATEGIES:
        bad_except(f"不支持的分块策略: {strategy}（可选 char/structure/semantic）")
    if strategy == "structure" and doc_type not in _STRUCTURE_DOC_TYPES:
        bad_except(
            f"章节分块暂不支持该文件类型: {doc_type}（仅 markdown/docx/excel），"
            "请改选通用分块或语义分块"
        )
    return strategy


# 支持的文件后缀 -> doc_type
_EXTENSION_DOC_TYPE = {
    "pdf": "pdf",
    "md": "markdown",
    "markdown": "markdown",
    "xlsx": "excel",
    "xls": "excel",
    "docx": "docx",
    "txt": "text",
    "csv": "csv",
}


def detect_doc_type(filename: str) -> str:
    """
    从文件名后缀推断 doc_type，不支持的类型抛业务异常（转为友好失败）。
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    doc_type = _EXTENSION_DOC_TYPE.get(ext)
    if doc_type is None:
        bad_except(f"不支持的文件类型: {filename or '(无扩展名)'}")
    return doc_type


# ---------- 各类型解析：返回 (full_text, page_ranges, sections) ----------
# page_ranges 为 [(page_number, start_char, end_char), ...]，非分页格式为空列表；
# sections 为 [(chapter_title | None, start_char, end_char), ...]（相对全文、
# 首标题前的前言段 title 为 None），仅结构化格式（markdown/docx/excel）产出，
# 其余格式为空列表（结构切分路径会退化为单节）。

# 行首 ATX 标题：1~6 个 # 后跟空白；7 个及以上不构成标题
_MD_HEADING_RE = re.compile(r"^(#{1,6})\s+(.*)$")


def _md_heading_title(raw: str) -> str | None:
    """清洗 ATX 标题文本：去首尾空白并剥掉可选的收尾 # 序列（须有空白前缀）。"""
    title = re.sub(r"\s+#+\s*$", "", raw.strip()).strip()
    return title or None


def _markdown_sections(text: str) -> list[tuple[str | None, int, int]]:
    """
    逐行扫描 markdown 标题产出节列表；维护 fenced code block（```/~~~）开合
    状态，代码块内的 # 行不构成节边界。标题行本身归入其所在节。
    """
    sections: list[tuple[str | None, int, int]] = []
    in_fence = False
    cur_title: str | None = None
    cur_start = 0
    offset = 0
    for line in text.splitlines(keepends=True):
        stripped = line.lstrip()
        if stripped.startswith("```") or stripped.startswith("~~~"):
            in_fence = not in_fence
        elif not in_fence:
            match = _MD_HEADING_RE.match(line)
            if match:
                # 关闭上一节（文件开头即标题时无前言段，不产出空节）
                if offset > cur_start:
                    sections.append((cur_title, cur_start, offset))
                cur_title = _md_heading_title(match.group(2))
                cur_start = offset
        offset += len(line)
    if len(text) > cur_start or not sections:
        sections.append((cur_title, cur_start, len(text)))
    return sections


def _parse_pdf(data: bytes) -> tuple[str, list[tuple[int, int, int]], list]:
    """按页解析 PDF 文本，记录每页在全文中的字符区间以便回填页码。"""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    page_ranges: list[tuple[int, int, int]] = []
    cursor = 0
    for page_number, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        start = cursor
        parts.append(page_text)
        cursor += len(page_text)
        page_ranges.append((page_number, start, cursor))
        # 页间以换行分隔（不计入某一页区间）
        parts.append("\n")
        cursor += 1
    full_text = "".join(parts)
    return full_text, page_ranges, []


# 标题样式名：英文 Word 为 "Heading N"，中文 Word 为 "标题 N"（空格可有可无）
_DOCX_HEADING_STYLE_RE = re.compile(r"^(Heading|标题)\s*\d+$")


def _parse_docx(data: bytes) -> tuple[str, list[tuple[int, int, int]], list]:
    """
    解析 docx，按段落拼接为纯文本；同时按 Heading/标题 样式段落累计游标产出
    节列表。无任何标题样式时退化为单节（chapter_title=None）。
    Word 无稳定页码，page_ranges 留空。
    """
    from docx import Document as DocxDocument

    doc = DocxDocument(io.BytesIO(data))
    texts: list[str] = []
    sections: list[tuple[str | None, int, int]] = []
    cur_title: str | None = None
    cur_start = 0
    cursor = 0
    for paragraph in doc.paragraphs:
        style_name = paragraph.style.name if paragraph.style is not None else ""
        if _DOCX_HEADING_STYLE_RE.match(style_name or ""):
            if cursor > cur_start:
                sections.append((cur_title, cur_start, cursor))
            cur_title = paragraph.text.strip() or None
            cur_start = cursor
        texts.append(paragraph.text)
        # 与 "\n".join 保持一致：每段贡献自身长度 + 1 个换行（末段的 +1 由
        # 收尾用 full_text 实际长度截断兜底）
        cursor += len(paragraph.text) + 1
    full_text = "\n".join(texts)
    if len(full_text) > cur_start or not sections:
        sections.append((cur_title, cur_start, len(full_text)))
    return full_text, [], sections


def _parse_xlsx(data: bytes) -> tuple[str, list[tuple[int, int, int]], list]:
    """解析 xlsx，逐表逐行把单元格以制表符/换行拼接为纯文本；每个工作表为一节。"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    sheet_marks: list[tuple[str | None, int]] = []  # (工作表名, 表头行号)
    for ws in wb.worksheets:
        sheet_marks.append((ws.title or None, len(lines)))
        lines.append(f"# {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            lines.append("\t".join(cells))
    wb.close()
    full_text = "\n".join(lines)
    # 行号 -> 全文字符偏移（每行贡献自身长度 + 1 个换行）
    line_offsets: list[int] = []
    cursor = 0
    for line in lines:
        line_offsets.append(cursor)
        cursor += len(line) + 1
    sections: list[tuple[str | None, int, int]] = []
    for i, (title, line_index) in enumerate(sheet_marks):
        start = line_offsets[line_index]
        end = (
            line_offsets[sheet_marks[i + 1][1]]
            if i + 1 < len(sheet_marks)
            else len(full_text)
        )
        sections.append((title, start, end))
    return full_text, [], sections


def _parse_markdown(data: bytes) -> tuple[str, list[tuple[int, int, int]], list]:
    """Markdown 按 UTF-8 文本处理（宽松解码），并扫描 ATX 标题产出节列表。"""
    text = data.decode("utf-8", errors="replace")
    return text, [], _markdown_sections(text)


def _parse_text(data: bytes) -> tuple[str, list[tuple[int, int, int]], list]:
    """纯文本按 UTF-8 宽松解码，无结构信号，不产出节列表。"""
    return data.decode("utf-8", errors="replace"), [], []


def _parse_csv(data: bytes) -> tuple[str, list[tuple[int, int, int]], list]:
    """解析 csv，单元格以制表符、行以换行拼接，与 xlsx 输出同构。"""
    import csv

    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = ["\t".join(row) for row in reader]
    return "\n".join(lines), [], []


_PARSERS = {
    "pdf": _parse_pdf,
    "docx": _parse_docx,
    "excel": _parse_xlsx,
    "markdown": _parse_markdown,
    "text": _parse_text,
    "csv": _parse_csv,
}


# ---------- 语义切分（semantic 策略） ----------
# 句子边界：换行与中英文句末标点（保留标点，空句剔除）
_SENTENCE_SPLIT_RE = re.compile(r"[^\n。！？!?;；]+(?:[。！？!?;；]|\n|$)")


def _split_sentences(text: str) -> list[tuple[str, int]]:
    """
    把文本切成句子序列，返回 [(sentence, start_offset), ...]（相对全文）。
    句子间可能存在未匹配间隙（如空行），父块构建时用原文切片保证内容一致。
    """
    return [
        (match.group(), match.start())
        for match in _SENTENCE_SPLIT_RE.finditer(text)
        if match.group().strip()
    ]


def _semantic_parent_texts(full_text: str) -> list[str]:
    """
    语义父块生成：句子级 embedding + 相邻相似度阈值断句 + 目标大小上限封口。

    步骤：句子切分 → embed_documents 批量取向量（一次调用）→ 相邻句 cosine
    相似度 → 低于 SEMANTIC_SIMILARITY_THRESHOLD 处切分；累积达
    SEMANTIC_TARGET_SIZE 也强制封口，避免超大块。产出父块原文切片
    （连续覆盖全文，不重叠不遗漏）。embedding 任何异常向上抛出，由
    ingest_file 降级处理。
    """
    sentences = _split_sentences(full_text)
    if len(sentences) <= 1:
        # 单句/空文档：无边界可切，整篇即一个父块
        return [full_text]

    client = EmbeddingFactory.get_client()
    embeddings = client.embed_documents([s for s, _ in sentences])
    similarities = [
        SimilarityUtils.cosine_similarity(
            np.asarray(embeddings[i]), np.asarray(embeddings[i + 1])
        )
        for i in range(len(sentences) - 1)
    ]

    groups: list[list[tuple[str, int]]] = []
    current: list[tuple[str, int]] = []
    current_len = 0
    for i, (sentence, _start) in enumerate(sentences):
        current.append((sentence, _start))
        current_len += len(sentence)
        is_last = i == len(sentences) - 1
        # 封口：非末句且（下一句相似度低于阈值 或 当前组已达目标大小）
        if not is_last and (
            similarities[i] < SEMANTIC_SIMILARITY_THRESHOLD
            or current_len >= SEMANTIC_TARGET_SIZE
        ):
            groups.append(current)
            current = []
            current_len = 0
    if current:
        groups.append(current)

    # 组内按首尾句子偏移取原文切片（句子间间隙保留在原文中，内容与全文一致）
    parents = []
    for group in groups:
        start = group[0][1]
        end = group[-1][1] + len(group[-1][0])
        parents.append(full_text[start:end])
    return parents


def _section_title_at(
    sections: list[tuple[str | None, int, int]], offset: int
) -> str | None:
    """按全文偏移定位所在章节标题；无结构信号或越界时返回 None。"""
    for title, start, end in sections:
        if start <= offset < end:
            return title
    return None


def _page_for_offset(
    page_ranges: list[tuple[int, int, int]], offset: int
) -> int | None:
    """把全文字符偏移映射到页码；无分页信息或越界时返回 None。"""
    if not page_ranges:
        return None
    for page_number, start, end in page_ranges:
        if start <= offset < end:
            return page_number
    # 落在末尾换行等边界时归到最后一页
    return page_ranges[-1][0]


def _find_offset(haystack: str, needle: str, cursor: int) -> int:
    """从 cursor 起定位子串偏移，找不到则退化为全局查找。"""
    idx = haystack.find(needle, cursor)
    if idx == -1:
        idx = haystack.find(needle)
    return max(idx, 0)


def _sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _build_parent_record(
    parent_text: str,
    chapter_title: str | None,
    p_start: int,
    p_end: int,
    page_ranges: list[tuple[int, int, int]],
    parent_index: int,
    leaf_index: int,
    child_splitter: RecursiveCharacterTextSplitter,
) -> tuple[dict[str, Any], int]:
    """
    父块内递归切叶块并产出父块记录（叶块偏移相对父块，页码经全文偏移回填）。
    返回 (parent_record, 下一个叶块全局顺序号)。
    """
    child_texts = child_splitter.split_text(parent_text)
    if not child_texts:
        child_texts = [parent_text]

    children: list[dict[str, Any]] = []
    child_cursor = 0
    for child_text in child_texts:
        c_start = _find_offset(parent_text, child_text, child_cursor)
        c_end = c_start + len(child_text)
        child_cursor = c_start + 1

        abs_start = p_start + c_start
        abs_end = p_start + c_end
        children.append(
            {
                "chunk_index": leaf_index,
                "content": child_text,
                "char_start": c_start,
                "char_end": c_end,
                "page_start": _page_for_offset(page_ranges, abs_start),
                "page_end": _page_for_offset(page_ranges, max(abs_end - 1, abs_start)),
                "chapter_title": chapter_title,
                "content_hash": _sha256(child_text),
            }
        )
        leaf_index += 1

    return (
        {
            "chunk_index": parent_index,
            "content": parent_text,
            "char_start": p_start,
            "char_end": p_end,
            "page_start": _page_for_offset(page_ranges, p_start),
            "page_end": _page_for_offset(page_ranges, max(p_end - 1, p_start)),
            "chapter_title": chapter_title,
            "content_hash": _sha256(parent_text),
            "children": children,
        },
        leaf_index,
    )


def ingest_file(
    filename: str, data: bytes, strategy: str = DEFAULT_CHUNK_STRATEGY
) -> dict[str, Any]:
    """
    把上传文件解析为文档字段 + 两级父子块树（不落库、不向量化）。

    strategy 为用户选择的切块策略（char/structure/semantic，见 CHUNK_STRATEGIES）：
    - char：全文直接字符递归切分，chapter_title 恒为 None（默认）；
    - structure：按章节/标题先切节，节内再字符切分，块携带 chapter_title；
      仅支持 markdown/docx/excel，其余类型抛业务异常；
    - semantic：句子级 embedding + 相邻相似度阈值切语义父块，父块内再字符切
      叶块，chapter_title 继承父块起点所在章节；embedding 服务异常时整体
      降级为 char 切分并记 warning。

    返回结构：
    {
      "doc_type": str,
      "content": str,               # 解析出的完整文本
      "content_hash": str,          # sha256(content)
      "effective_strategy": str,    # 生效策略（auto 归一化后为 char/structure）
      "parents": [                  # level 1 父块
        {
          "chunk_index": int,       # 父块在 level 1 内的顺序号
          "content": str,
          "char_start": int,        # 相对全文
          "char_end": int,
          "page_start": int | None,
          "page_end": int | None,
          "chapter_title": str | None,   # 所属章节标题（char 路径恒 None）
          "content_hash": str,
          "children": [             # level 0 叶块
            {
              "chunk_index": int,   # 叶块在 level 0 内的全局顺序号
              "content": str,
              "char_start": int,    # 相对父块
              "char_end": int,
              "page_start": int | None,
              "page_end": int | None,
              "chapter_title": str | None,  # 继承父块
              "content_hash": str,
            }, ...
          ],
        }, ...
      ],
    }
    """
    doc_type = detect_doc_type(filename)
    strategy = validate_chunk_strategy(strategy, doc_type)
    parser = _PARSERS[doc_type]
    # 解析异常统一转为友好业务失败：扩展名匹配但内容损坏/伪造时不穿透为 500
    try:
        full_text, page_ranges, sections = parser(data)
    except BadException:
        raise
    except Exception as exc:
        logger.warning(f"文件解析失败 [{filename}] doc_type={doc_type}: {exc!r}")
        bad_except(f"文件已损坏或与扩展名不符: {filename}")
    content_hash = _sha256(full_text)

    parent_splitter = RecursiveCharacterTextSplitter(
        chunk_size=PARENT_CHUNK_SIZE, chunk_overlap=0
    )
    child_splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHILD_CHUNK_SIZE, chunk_overlap=CHILD_CHUNK_OVERLAP
    )

    parents: list[dict[str, Any]] = []
    parent_index = 0  # level 1 全局顺序号（跨节连续）
    leaf_index = 0  # level 0 全局顺序号（跨父块连续）

    if strategy == "semantic":
        # 语义分支：全文统一语义切父块，不约束章节边界；embedding 任何异常
        # 整体降级为 char（生效策略随之记 char），保证上传不中断
        try:
            parent_texts = _semantic_parent_texts(full_text)
            effective_strategy = "semantic"
        except Exception as exc:
            logger.warning(f"语义分块失败，降级为字符分块: {exc!r}")
            parent_texts = parent_splitter.split_text(full_text)
            if not parent_texts:
                parent_texts = [full_text]
            effective_strategy = "char"
        cursor = 0
        for parent_text in parent_texts:
            p_start = _find_offset(full_text, parent_text, cursor)
            p_end = p_start + len(parent_text)
            cursor = p_start + 1
            chapter_title = _section_title_at(sections, p_start)
            parent, leaf_index = _build_parent_record(
                parent_text,
                chapter_title,
                p_start,
                p_end,
                page_ranges,
                parent_index,
                leaf_index,
                child_splitter,
            )
            parents.append(parent)
            parent_index += 1
    else:
        # char/structure 分支：structure 且格式支持时按章节先切节；char 或
        # 无结构格式统一退化为覆盖全文的单节，与历史字符切分逐字节一致
        use_structure = strategy == "structure" and doc_type in _STRUCTURE_DOC_TYPES
        effective_strategy = "structure" if use_structure else "char"
        if not use_structure or not sections:
            sections = [(None, 0, len(full_text))]

        for chapter_title, sec_start, sec_end in sections:
            section_text = full_text[sec_start:sec_end]
            parent_texts = parent_splitter.split_text(section_text)
            # 空文档或纯空白时也产出一个覆盖本节的父块，保证结构完整
            if not parent_texts:
                parent_texts = [section_text]

            sec_cursor = 0
            for parent_text in parent_texts:
                p_local = _find_offset(section_text, parent_text, sec_cursor)
                p_start = sec_start + p_local
                p_end = p_start + len(parent_text)
                sec_cursor = p_local + 1

                parent, leaf_index = _build_parent_record(
                    parent_text,
                    chapter_title,
                    p_start,
                    p_end,
                    page_ranges,
                    parent_index,
                    leaf_index,
                    child_splitter,
                )
                parents.append(parent)
                parent_index += 1

    return {
        "doc_type": doc_type,
        "content": full_text,
        "content_hash": content_hash,
        "effective_strategy": effective_strategy,
        "parents": parents,
    }
